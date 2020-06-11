import urlparse
import MySQLdb
import datetime
import sys
import time
import re
from vtvspider import VTVSpider, get_nodes, extract_data, get_utc_time
from vtvspider import get_tzinfo
from scrapy.selector import Selector
from scrapy.http import Request
import unicodedata
from StringUtil import cleanString
from difflib import SequenceMatcher
import codecs
from openpyxl import load_workbook
from vtv_task import VtvTask, vtv_task_main

PLAYERS_PERMUTATIONS = {}
TEAMS_IN_DB        = {}
TEAMS_XLS = {}
SPACE = ' '
THRESHOLD   = 0.70


def permutations(iterable, r=None):
    # permutations('ABCD', 2) --> AB AC AD BA BC BD CA CB CD DA DB DC
    # permutations(range(3)) --> 012 021 102 120 201 210
    pool = tuple(iterable)
    n = len(pool)
    if r is None:
        r = n 

    if r > n:
        return
    indices = range(n)
    cycles = range(n, n-r, -1) 
    yield tuple(pool[i] for i in indices[:r])
    while n:
        for i in reversed(range(r)):
            cycles[i] -= 1
            if cycles[i] == 0:
                indices[i:] = indices[i+1:] + indices[i:i+1]
                cycles[i] = n - i 
            else:
                j = cycles[i]
                indices[i], indices[-j] = indices[-j], indices[i]
                yield tuple(pool[i] for i in indices[:r])
                break
        else:
            return

def get_player_permutations(existing_players):
    for existing_player in existing_players:
        if SPACE in existing_player:
            existing_player = cleanString(existing_player)
            existing_player_list = existing_player.split(SPACE)
            if len(existing_player_list) <= 5:
                existing_player_set = [SPACE.join(i) for i in \
                                      permutations(existing_player_list)]
            else:
                existing_player_set = [existing_player]

        else:
            existing_player_set = [existing_player]

        PLAYERS_PERMUTATIONS[existing_player] = existing_player_set


def find_players_to_add(db_players, tou_players):
    titles_present = {}
    mstly_matched_titles = {}

    for player in tou_players:
        player = player
        if player in db_players:
            titles_present[player] = (player, 'EXACTLY PRESENT')
            continue
        check_flag = False
        seqmatcher = SequenceMatcher(None, player)

        for existing_player in db_players:
            if SPACE in player and SPACE in existing_player:
                existing_player_set = PLAYERS_PERMUTATIONS.\
                                      get(existing_player, [existing_player])
            else:
                existing_player_set = [existing_player]

            for ex_player in existing_player_set:
                seqmatcher.set_seq2(ex_player)
                if seqmatcher.ratio() >= THRESHOLD:
                    mstly_matched_titles[player] = existing_player
                    titles_present[player] = (player, 'MOSTLY PRESENT')
                    check_flag = True
                    break

            if check_flag: break
        else:
            titles_present[player] = (player, 'NOT PRESENT')

    mostly_present = [player for player, details in titles_present.items() \
                     if details[1] == 'MOSTLY PRESENT']
    exactly_present = [player for player, details in titles_present.items() \
                     if details[1] == 'EXACTLY PRESENT']
    not_present = [player for player, details in titles_present.items() \
                     if details[1] == 'NOT PRESENT']
    return exactly_present, mostly_present, not_present, mstly_matched_titles

class XlsxReader(VtvTask):

    def __init__(self):
        VtvTask.__init__(self)
        self.conn = MySQLdb.connect(host="10.4.18.183", user="root", db="SPORTSDB", charset='utf8', use_unicode=True)
        self.cursor = self.conn.cursor()
        self.players_file = open('teams_variants_file.txt', 'a+')

    def parse_xls(self):
        self.get_db_teams()
        wb = load_workbook('Translations.xlsx', read_only=True)
        sheets = wb.sheetnames
        for sht in sheets:
            if 'Leagues' in sht: continue
            ws = wb.get_sheet_by_name(sht)
            for w in ws:
                tou, fr, ita, ger, spa = w[0].value, w[1].value, w[2].value, w[3].value, w[4].value
                TEAMS_XLS.update({tou.lower(): (fr, ita, ger, spa)})
                
            get_player_permutations(self.TEAMS_IN_DB.keys())

            for tmi in TEAMS_XLS.keys():
                
                _team = tmi.strip()
                _team_id   = self.TEAMS_IN_DB.get(_team.lower())
                if not _team_id:
                    ex, x, y, z = find_players_to_add([_team], self.TEAMS_IN_DB.keys()) 
                    try:
                        _team_id = self.TEAMS_IN_DB[x[0].lower()]
                    except: 
                        _team_id = ''
                if _team_id:
                    try:
                        self.parse_player_insertion(_team_id, TEAMS_XLS[_team.lower()])
                    except:
                        self.players_file.write('%s<>%s\n'%(_team, TEAMS_XLS[_team.lower()]))
                else:
                    self.players_file.write('%s<>%s\n'%(_team, TEAMS_XLS[_team.lower()]))

    def parse_player_insertion(self, player_id, player):
        en_id = player_id
        fr, ita, ger, spa = player

        qry = 'insert into sports_titles_regions (entity_id, entity_type, title, aka, short_title, iso, description, created_at, modified_at) values (%s, %s, %s, "", "", %s, "", now(), now()) on duplicate key update modified_at = now()'
        vals = (en_id, 'team', fr, 'FRA')
        self.cursor.execute(qry, vals)

        qry = 'insert into sports_titles_regions (entity_id, entity_type, title, aka, short_title, iso, description, created_at, modified_at) values (%s, %s, %s, "", "", %s, "", now(), now()) on duplicate key update modified_at = now()'
        vals = (en_id, 'team', ita, 'ITA')
        self.cursor.execute(qry, vals)

        qry = 'insert into sports_titles_regions (entity_id, entity_type, title, aka, short_title, iso, description, created_at, modified_at) values (%s, %s, %s, "", "", %s, "", now(), now()) on duplicate key update modified_at = now()'
        vals = (en_id, 'team', ger, 'DEU')
        self.cursor.execute(qry, vals)

        qry = 'insert into sports_titles_regions (entity_id, entity_type, title, aka, short_title, iso, description, created_at, modified_at) values (%s, %s, %s, "", "", %s, "", now(), now()) on duplicate key update modified_at = now()'
        vals = (en_id, 'team', spa, 'ESP')
        self.cursor.execute(qry, vals)



    def get_db_teams(self):
        query = 'select id, title from sports_participants where game = "soccer"'
        self.cursor.execute(query)
        teams = self.cursor.fetchall()

        self.TEAMS_IN_DB = {}
        for tm in teams:
            tm_id, tm_title = tm
            if tm_id:
                tm_title = tm_title.strip().lower()
                self.TEAMS_IN_DB[tm_title] = int(tm_id)
    def run_main(self):
        self.parse_xls()

if __name__ == '__main__':
    vtv_task_main(XlsxReader)
    sys.exit( 0 ) 

